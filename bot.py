"""
Telegram-бот для просмотра реестра задач СЦ СОЛВО.
Поиск по номеру задачи и по ключевому слову.
"""

import os
import logging
from openpyxl import load_workbook
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    CallbackQueryHandler, ContextTypes, filters
)

# ─── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    format="%(asctime)s | %(levelname)s | %(message)s",
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# ─── Config ───────────────────────────────────────────────────────────────────
BOT_TOKEN  = os.environ.get("BOT_TOKEN", "PASTE_YOUR_TOKEN_HERE")
EXCEL_PATH = os.environ.get("EXCEL_PATH", "registry.xlsx")

# ─── Status emoji map ─────────────────────────────────────────────────────────
STATUS_EMOJI = {
    "Установлено":          "✅",
    "Выполнено":            "✅",
    "Не СОЛВО":             "☑️",
    "ОПЭ":                  "🔄",
    "Тестирование":         "🧪",
    "На приемке":           "📥",
    "В работе":             "⚙️",
    "Разработка":           "👨‍💻",
    "Аналитика":            "🔍",
    "Оценка":               "📊",
    "Ожидает рассмотрения": "⏳",
    "Приостановлена":       "⏸️",
}

PRIORITY_EMOJI = {
    "Высокий": "🔴",
    "Средний":  "🟡",
    "Низкий":   "🟢",
}

# ─── Data loader ─────────────────────────────────────────────────────────────
def load_tasks() -> list[dict]:
    """Читает реестр из Excel и возвращает список задач."""
    try:
        wb = load_workbook(EXCEL_PATH, data_only=True)
        ws = wb["📋 Реестр"]
    except Exception as e:
        logger.error(f"Не удалось открыть файл: {e}")
        return []

    tasks = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        num = row[0]   # col A: №
        # Пропускаем пустые строки и строки-разделители категорий
        if not isinstance(num, int):
            continue

        task = {
            "num":          num,
            "ticket":       str(row[1] or "").strip(),
            "name":         str(row[2] or "").strip(),
            "desc":         str(row[3] or "").strip(),
            "deadline":     str(row[4] or "").strip(),
            "status_sc":    str(row[5] or "").strip(),
            "status_solvo": str(row[6] or "").strip(),
            "release":      str(row[7] or "").strip(),
            "hours":        row[8] if row[8] else 0,
            "agreed":       str(row[9] or "").strip(),
            "category":     str(row[10] or "").strip(),
            "priority":     str(row[12] or "").strip(),
            "notes":        str(row[13] or "").strip() if len(row) > 13 else "",
        }
        tasks.append(task)

    logger.info(f"Загружено задач: {len(tasks)}")
    return tasks


def format_task_card(t: dict) -> str:
    """Форматирует карточку задачи для отправки в Telegram."""
    sc_ico   = STATUS_EMOJI.get(t["status_sc"],    "▪️")
    sv_ico   = STATUS_EMOJI.get(t["status_solvo"], "▪️")
    # Вычисляем приоритет сами если не прочитался из файла
    if not t["priority"]:
        sv = t["status_solvo"]
        if sv in ("Тестирование","Разработка","На приемке"):
            t["priority"] = "Высокий"
        elif sv in ("В работе","Аналитика","Оценка","ОПЭ"):
            t["priority"] = "Средний"
        else:
            t["priority"] = "Низкий"
    pr_ico   = PRIORITY_EMOJI.get(t["priority"],   "▪️")
    agreed   = "✅ Да" if t["agreed"] == "Да" else "❌ Нет"
    hours    = f"{int(t['hours'])} ч/ч" if t["hours"] else "—"
    release  = t["release"] if t["release"] else "—"
    notes    = f"\n💬 *Примечание:* {t['notes']}" if t["notes"] else ""
    ticket   = f"\n🎫 *Заявка:* `{t['ticket']}`" if t["ticket"] else ""

    return (
        f"📋 *Задача #{t['num']}*\n"
        f"━━━━━━━━━━━━━━━━━━━\n"
        f"*{t['name']}*\n\n"
        f"📝 {t['desc']}\n"
        f"{ticket}\n\n"
        f"📂 *Категория:* {t['category']}\n"
        f"📅 *Срок:* {t['deadline']}\n"
        f"🚀 *Релиз:* {release}\n\n"
        f"{sc_ico} *Статус СЦ:* {t['status_sc']}\n"
        f"{sv_ico} *Статус СОЛВО:* {t['status_solvo']}\n\n"
        f"⏱ *Оценка:* {hours}\n"
        f"💰 *Оценена:* {agreed}\n"
        f"{pr_ico} *Приоритет:* {t['priority']}"
        f"{notes}"
    )


def format_summary() -> str:
    """Общая сводка по реестру."""
    tasks = load_tasks()
    if not tasks:
        return "⚠️ Не удалось загрузить реестр."

    total  = len(tasks)
    done   = sum(1 for t in tasks if t["status_solvo"] in ("Установлено","Не СОЛВО","Выполнено"))
    active = sum(1 for t in tasks if t["status_solvo"] in ("В работе","Тестирование","Разработка","ОПЭ","На приемке"))
    analyt = sum(1 for t in tasks if t["status_solvo"] in ("Аналитика","Оценка","Ожидает рассмотрения"))
    paused = sum(1 for t in tasks if t["status_solvo"] == "Приостановлена")
    hours  = sum(int(t["hours"]) for t in tasks if t["hours"])
    pct    = round(done / total * 100) if total else 0

    # По категориям
    cats = {}
    for t in tasks:
        c = t["category"]
        if c not in cats:
            cats[c] = {"total":0,"done":0}
        cats[c]["total"] += 1
        if t["status_solvo"] in ("Установлено","Не СОЛВО","Выполнено"):
            cats[c]["done"] += 1

    cat_lines = ""
    for cat, v in cats.items():
        bar = "█" * (v["done"] * 5 // v["total"]) + "░" * (5 - v["done"] * 5 // v["total"])
        cat_lines += f"  {bar} {cat[:30]}: {v['done']}/{v['total']}\n"

    return (
        f"📊 *СВОДКА ПО РЕЕСТРУ*\n"
        f"━━━━━━━━━━━━━━━━━━━\n"
        f"📌 Всего задач: *{total}*\n"
        f"✅ Выполнено: *{done}* ({pct}%)\n"
        f"⚙️ В работе: *{active}*\n"
        f"🔍 Аналитика/Оценка: *{analyt}*\n"
        f"⏸️ Приостановлено: *{paused}*\n"
        f"⏱ Суммарная оценка: *{hours} ч/ч*\n\n"
        f"*По разделам:*\n"
        f"`{cat_lines}`"
    )


# ─── Handlers ─────────────────────────────────────────────────────────────────

async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = (
        "👋 Привет! Я бот реестра задач *СЦ СОЛВО*.\n\n"
        "Что умею:\n"
        "🔢 `/task 14` — карточка задачи по номеру\n"
        "🔍 `/find букинг` — поиск по ключевому слову\n"
        "📊 `/summary` — общая сводка по реестру\n"
        "📂 `/category терминал` — задачи по разделу\n"
        "🚦 `/status тестирование` — задачи по статусу\n\n"
        "Или просто напиши номер или слово — найду сам!"
    )
    keyboard = [
        [InlineKeyboardButton("📊 Сводка", callback_data="summary")],
        [
            InlineKeyboardButton("🔍 Поиск по слову", callback_data="hint_find"),
            InlineKeyboardButton("🔢 По номеру",      callback_data="hint_num"),
        ],
    ]
    await update.message.reply_text(
        text, parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def cmd_task(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Поиск задачи по номеру: /task 14"""
    if not context.args:
        await update.message.reply_text("Укажи номер задачи. Пример: `/task 14`", parse_mode="Markdown")
        return

    try:
        num = int(context.args[0])
    except ValueError:
        await update.message.reply_text("Номер должен быть числом. Пример: `/task 14`", parse_mode="Markdown")
        return

    tasks = load_tasks()
    found = [t for t in tasks if t["num"] == num]

    if not found:
        await update.message.reply_text(f"❌ Задача #{num} не найдена в реестре.")
        return

    await update.message.reply_text(format_task_card(found[0]), parse_mode="Markdown")


async def cmd_find(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Поиск по ключевому слову: /find букинг"""
    if not context.args:
        await update.message.reply_text("Укажи слово для поиска. Пример: `/find растарка`", parse_mode="Markdown")
        return

    query = " ".join(context.args).lower().strip()
    tasks = load_tasks()

    found = [
        t for t in tasks
        if query in t["name"].lower()
        or query in t["desc"].lower()
        or query in t["ticket"].lower()
        or query in t["category"].lower()
    ]

    if not found:
        await update.message.reply_text(f"❌ По запросу «{query}» ничего не найдено.")
        return

    if len(found) == 1:
        await update.message.reply_text(format_task_card(found[0]), parse_mode="Markdown")
        return

    # Несколько результатов — показываем список с кнопками
    text = f"🔍 Найдено задач: *{len(found)}* по запросу «{query}»\n\nВыбери задачу:\n"
    keyboard = []
    for t in found[:10]:  # максимум 10
        sc_ico = STATUS_EMOJI.get(t["status_sc"], "▪️")
        label  = f"#{t['num']} {t['name'][:35]} {sc_ico}"
        keyboard.append([InlineKeyboardButton(label, callback_data=f"task_{t['num']}")])

    await update.message.reply_text(
        text, parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def cmd_summary(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(format_summary(), parse_mode="Markdown")


async def cmd_category(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Задачи по разделу: /category терминал"""
    if not context.args:
        await update.message.reply_text(
            "Укажи раздел. Примеры:\n`/category терминал`\n`/category судовые`\n`/category жд`",
            parse_mode="Markdown"
        )
        return

    ALIASES = {
        "терминал":           "терминальные",
        "терминальные":       "терминальные",
        "судовые":            "судовых",
        "судовые документы":  "судовых",
        "таможня":            "таможенных",
        "таможен":            "таможенных",
        "таможенные":         "таможенных",
        "экспорт":            "экспортных",
        "экспортные":         "экспортных",
        "жд":                 "жд",
        "жд грузы":           "жд",
        "ж/д":                "жд",
        "железная дорога":    "жд",
        "железнодорожн":      "жд",
    }
    query = " ".join(context.args).lower()
    search = ALIASES.get(query, query)
    tasks = load_tasks()
    found = [t for t in tasks if search in t["category"].lower()]

    if not found:
        await update.message.reply_text(f"❌ Раздел «{query}» не найден.")
        return

    cat_name = found[0]["category"]
    done     = sum(1 for t in found if t["status_solvo"] in ("Установлено","Не СОЛВО","Выполнено"))
    lines    = [f"📂 *{cat_name}* — {len(found)} задач ({done} выполнено)\n"]

    for t in found:
        sc_ico = STATUS_EMOJI.get(t["status_sc"],    "▪️")
        sv_ico = STATUS_EMOJI.get(t["status_solvo"], "▪️")
        lines.append(f"{sv_ico} *#{t['num']}* {t['name']}\n   СЦ: {t['status_sc']} | СОЛВО: {t['status_solvo']}")

    keyboard = [
        [InlineKeyboardButton(f"#{t['num']} подробнее", callback_data=f"task_{t['num']}")]
        for t in found[:8]
    ]
    await update.message.reply_text(
        "\n".join(lines), parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def cmd_status(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Задачи по статусу: /status тестирование"""
    if not context.args:
        await update.message.reply_text(
            "Укажи статус. Пример: `/status тестирование`",
            parse_mode="Markdown"
        )
        return

    query = " ".join(context.args).lower()
    tasks = load_tasks()
    found = [
        t for t in tasks
        if query in t["status_sc"].lower() or query in t["status_solvo"].lower()
    ]

    if not found:
        await update.message.reply_text(f"❌ Задач со статусом «{query}» не найдено.")
        return

    lines = [f"🚦 Задачи со статусом «{query}» — *{len(found)} шт.*\n"]
    for t in found:
        sv_ico = STATUS_EMOJI.get(t["status_solvo"], "▪️")
        lines.append(f"{sv_ico} *#{t['num']}* {t['name']}\n   СОЛВО: {t['status_solvo']} | Срок: {t['deadline']}")

    keyboard = [
        [InlineKeyboardButton(f"#{t['num']} подробнее", callback_data=f"task_{t['num']}")]
        for t in found[:8]
    ]
    await update.message.reply_text(
        "\n".join(lines), parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Свободный ввод: число → поиск по номеру, текст → поиск по слову."""
    text = update.message.text.strip()

    # Если число — ищем по номеру
    if text.isdigit():
        context.args = [text]
        await cmd_task(update, context)
        return

    # Иначе — поиск по ключевому слову
    context.args = text.split()
    await cmd_find(update, context)


async def handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка нажатий на inline-кнопки."""
    query = update.callback_query
    await query.answer()
    data = query.data

    if data == "summary":
        await query.message.reply_text(format_summary(), parse_mode="Markdown")

    elif data == "hint_find":
        await query.message.reply_text(
            "Напиши ключевое слово, например:\n`букинг`, `растарка`, `жд`, `таможня`",
            parse_mode="Markdown"
        )
    elif data == "hint_num":
        await query.message.reply_text(
            "Напиши номер задачи, например: `14` или `/task 14`",
            parse_mode="Markdown"
        )
    elif data.startswith("task_"):
        num = int(data.split("_")[1])
        tasks = load_tasks()
        found = [t for t in tasks if t["num"] == num]
        if found:
            await query.message.reply_text(format_task_card(found[0]), parse_mode="Markdown")


async def cmd_registry(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отправить файл реестра: /registry"""
    import os
    if not os.path.exists(EXCEL_PATH):
        await update.message.reply_text("Файл реестра не найден на сервере.")
        return
    tasks = load_tasks()
    caption = f"Реестр задач автоматизации СЦ СОЛВО\nВсего задач: {len(tasks)}"
    await update.message.reply_text("Отправляю файл реестра...")
    with open(EXCEL_PATH, "rb") as f:
        await update.message.reply_document(
            document=f,
            filename="Реестр_Автоматизация_СЦ_СОЛВО_2026.xlsx",
            caption=caption,
        )

# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    if BOT_TOKEN == "PASTE_YOUR_TOKEN_HERE":
        print("❌ Укажи BOT_TOKEN в переменной окружения или в коде!")
        return

    app = Application.builder().token(BOT_TOKEN).build()

    app.add_handler(CommandHandler("start",    cmd_start))
    app.add_handler(CommandHandler("task",     cmd_task))
    app.add_handler(CommandHandler("find",     cmd_find))
    app.add_handler(CommandHandler("summary",  cmd_summary))
    app.add_handler(CommandHandler("category", cmd_category))
    app.add_handler(CommandHandler("status",    cmd_status))
    app.add_handler(CommandHandler("registry",  cmd_registry))
    app.add_handler(CallbackQueryHandler(handle_callback))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))

    logger.info("Бот запущен. Жду сообщений...")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
