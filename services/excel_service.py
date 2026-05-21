import openpyxl

# ─────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────

FILE_PATH = "registry.xlsx"

# ─────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────

def safe_str(value):

    if value is None:
        return ""

    return str(value).strip()

def safe_float(value):

    try:

        if value is None:
            return 0

        return float(value)

    except:

        return 0

# ─────────────────────────────────────────────────────
# LOAD TASKS
# ─────────────────────────────────────────────────────

def load_tasks():

    workbook = openpyxl.load_workbook(
        FILE_PATH,
        data_only=True
    )

    tasks = []

    # ─────────────────────────────────────────────
    # ИЩЕМ ЛИСТ РЕЕСТРА
    # ─────────────────────────────────────────────

    sheet = None

    for sheet_name in workbook.sheetnames:

        if "реестр" in sheet_name.lower():

            sheet = workbook[sheet_name]

            break

    if sheet is None:

        print("❌ Лист реестра не найден")

        return []

    print(f"✅ Используется лист: {sheet.title}")

    # ─────────────────────────────────────────────
    # ЧИТАЕМ СТРОКИ
    # ─────────────────────────────────────────────

    for row in sheet.iter_rows(
        min_row=2,
        values_only=True
    ):

        if not row:
            continue

        # ─────────────────────────────────────────
        # НОМЕР
        # ─────────────────────────────────────────

        num = safe_str(row[0])

        if not num:
            continue

        # Пропускаем служебные строки
        if num.lower().strip() in [

            "категория",
            "category",

        ]:
            continue

        # ─────────────────────────────────────────
        # НАЗВАНИЕ
        # ─────────────────────────────────────────

        name = safe_str(row[2])

        if not name:
            continue

        normalized_name = (

            name
            .lower()
            .replace(":", "")
            .strip()

        )

        # Пропускаем служебные строки
        if normalized_name in [

            "категория",
            "categories",

        ]:
            continue

        # ─────────────────────────────────────────
        # КАТЕГОРИЯ
        # ─────────────────────────────────────────

        category = safe_str(row[10])

        normalized_category = (

            category
            .lower()
            .replace(":", "")
            .strip()

        )

        # Пропускаем мусорные строки
        if normalized_category in [

            "категория",
            "categories",

        ]:
            continue

        # ─────────────────────────────────────────
        # TASK
        # ─────────────────────────────────────────

        task = {

            # A
            "num": num,

            # B
            "ticket": safe_str(row[1]),

            # C
            "name": name,

            # D
            "desc": safe_str(row[3]),

            # E
            "deadline": safe_str(row[4]),

            # F
            "status_sc": safe_str(row[5]),

            # G
            "status_solvo": safe_str(row[6]),

            # H
            "release": safe_str(row[7]),

            # I
            "hours": safe_float(row[8]),

            # J
            "agreed": safe_str(row[9]),

            # K
            "category": category,

            # L
            "priority": safe_str(row[11]),

            # M
            "notes": safe_str(row[12]),
        }

        tasks.append(task)

    print(f"📊 Загружено задач: {len(tasks)}")

    return tasks

# ─────────────────────────────────────────────────────
# GET TASK
# ─────────────────────────────────────────────────────

def get_task(task_num):

    tasks = load_tasks()

    for task in tasks:

        if str(task["num"]) == str(task_num):

            return task

    return None

# ─────────────────────────────────────────────────────
# SEARCH TASKS
# ─────────────────────────────────────────────────────

def search_tasks(query):

    query = query.lower()

    tasks = load_tasks()

    found = []

    for task in tasks:

        text = (

            f"{task['name']} "
            f"{task['desc']} "
            f"{task['ticket']} "
            f"{task['category']} "
            f"{task['release']} "
            f"{task['status_solvo']} "
            f"{task['priority']}"

        ).lower()

        if query in text:

            found.append(task)

    return found

# ─────────────────────────────────────────────────────
# FILTER BY STATUS
# ─────────────────────────────────────────────────────

def get_tasks_by_status(status):

    tasks = load_tasks()

    return [

        t for t in tasks

        if t["status_solvo"] == status

    ]

# ─────────────────────────────────────────────────────
# FILTER BY CATEGORY
# ─────────────────────────────────────────────────────

def get_tasks_by_category(category):

    tasks = load_tasks()

    return [

        t for t in tasks

        if t["category"] == category

    ]

# ─────────────────────────────────────────────────────
# ADVANCED FILTERS
# ─────────────────────────────────────────────────────

def filter_tasks(
    category=None,
    status=None,
    release=None,
    priority=None
):

    tasks = load_tasks()

    if category:

        tasks = [

            t for t in tasks

            if t["category"] == category

        ]

    if status:

        tasks = [

            t for t in tasks

            if t["status_solvo"] == status

        ]

    if release:

        tasks = [

            t for t in tasks

            if t["release"] == release

        ]

    if priority:

        tasks = [

            t for t in tasks

            if t["priority"] == priority

        ]

    return tasks